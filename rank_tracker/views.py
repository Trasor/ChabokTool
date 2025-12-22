from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.db import transaction
from .models import RankProject, RankKeyword, RankHistory
from .forms import ProjectCreateForm, KeywordAddForm, KeywordImportForm
from .services import RankService
from billing.models import UserCredit
import pandas as pd
import openpyxl
from openpyxl.styles import Font
import logging

logger = logging.getLogger(__name__)


@login_required
def project_list(request):
    """لیست پروژه‌های کاربر"""
    projects = RankProject.objects.filter(user=request.user, is_active=True)

    context = {
        'projects': projects
    }
    return render(request, 'rank_tracker/project_list.html', context)


@login_required
def project_create(request):
    """ساخت پروژه جدید"""
    if request.method == 'POST':
        form = ProjectCreateForm(request.POST)
        if form.is_valid():
            # محاسبه هزینه
            capacity = form.cleaned_data['keyword_capacity']
            cost = (capacity // 100) * 500

            # بررسی موجودی
            user_credit, created = UserCredit.objects.get_or_create(user=request.user)

            if user_credit.balance < cost:
                messages.error(
                    request,
                    f'❌ موجودی کافی نیست! نیاز: {cost} کردیت | موجودی فعلی: {user_credit.balance} کردیت'
                )
                return redirect('transactions_list')

            # کسر کردیت و ساخت پروژه
            with transaction.atomic():
                user_credit.balance -= cost
                user_credit.save()

                project = form.save(commit=False)
                project.user = request.user
                project.save()

                logger.info(f"✓ Project '{project.project_name}' created by {request.user.username} - Cost: {cost} credits")
                messages.success(request, f'✅ پروژه با موفقیت ساخته شد! ({cost} کردیت کسر شد)')
                return redirect('rank_tracker:keyword_add', project_id=project.id)

    else:
        form = ProjectCreateForm()

    # دریافت موجودی کاربر
    user_credit = UserCredit.objects.filter(user=request.user).first()
    balance = user_credit.balance if user_credit else 0

    context = {
        'form': form,
        'balance': balance
    }
    return render(request, 'rank_tracker/project_create.html', context)


@login_required
def project_detail(request, project_id):
    """جزئیات پروژه و جدول کلمات کلیدی"""
    project = get_object_or_404(RankProject, id=project_id, user=request.user)

    # بررسی و reset شمارنده ماهانه
    project.check_and_reset_monthly_counter()

    keywords = project.keywords.all().order_by('keyword')

    context = {
        'project': project,
        'keywords': keywords
    }
    return render(request, 'rank_tracker/project_detail.html', context)


@login_required
def project_delete(request, project_id):
    """حذف پروژه"""
    project = get_object_or_404(RankProject, id=project_id, user=request.user)

    if request.method == 'POST':
        project_name = project.project_name
        project.is_active = False  # Soft delete
        project.save()

        messages.success(request, f'✅ پروژه "{project_name}" حذف شد.')
        return redirect('rank_tracker:project_list')

    return render(request, 'rank_tracker/project_delete_confirm.html', {'project': project})


@login_required
def keyword_add(request, project_id):
    """اضافه کردن کلمات کلیدی (دستی یا Excel)"""
    project = get_object_or_404(RankProject, id=project_id, user=request.user)

    if request.method == 'POST':
        # تشخیص نوع فرم
        if 'keywords' in request.POST:
            # فرم دستی
            form = KeywordAddForm(request.POST)
            if form.is_valid():
                keywords_list = form.cleaned_data['keywords']
                return _process_keywords(request, project, keywords_list)

        elif 'file' in request.FILES:
            # فرم آپلود
            form = KeywordImportForm(request.POST, request.FILES)
            if form.is_valid():
                file = form.cleaned_data['file']
                try:
                    # خواندن Excel
                    df = pd.read_excel(file, header=0)

                    # بررسی ستون keyword
                    if 'keyword' not in df.columns:
                        messages.error(request, '❌ فایل باید ستونی با نام "keyword" داشته باشد.')
                        return redirect('rank_tracker:keyword_add', project_id=project.id)

                    # استخراج کلمات
                    keywords_list = df['keyword'].dropna().astype(str).str.strip().tolist()
                    keywords_list = [kw for kw in keywords_list if kw]  # حذف خالی‌ها

                    if not keywords_list:
                        messages.error(request, '❌ فایل خالی است!')
                        return redirect('rank_tracker:keyword_add', project_id=project.id)

                    return _process_keywords(request, project, keywords_list)

                except Exception as e:
                    logger.error(f"✗ Excel import error: {str(e)}")
                    messages.error(request, f'❌ خطا در خواندن فایل: {str(e)}')

    else:
        manual_form = KeywordAddForm()
        import_form = KeywordImportForm()

    context = {
        'project': project,
        'manual_form': manual_form if request.method == 'GET' else KeywordAddForm(),
        'import_form': import_form if request.method == 'GET' else KeywordImportForm()
    }
    return render(request, 'rank_tracker/keyword_add.html', context)


def _process_keywords(request, project, keywords_list):
    """پردازش و ذخیره کلمات کلیدی"""
    # بررسی ظرفیت
    current_count = project.keywords_count
    new_count = len(keywords_list)
    total = current_count + new_count

    if total > project.keyword_capacity:
        messages.error(
            request,
            f'❌ ظرفیت تکمیل است! فعلی: {current_count} | جدید: {new_count} | '
            f'ظرفیت: {project.keyword_capacity} | مازاد: {total - project.keyword_capacity}'
        )
        return redirect('rank_tracker:keyword_add', project_id=project.id)

    # ذخیره کلمات
    added = 0
    duplicates = 0

    for keyword in keywords_list:
        if not keyword.strip():
            continue

        _, created = RankKeyword.objects.get_or_create(
            project=project,
            keyword=keyword.strip()
        )

        if created:
            added += 1
        else:
            duplicates += 1

    if added > 0:
        messages.success(request, f'✅ {added} کلمه کلیدی اضافه شد!')
    if duplicates > 0:
        messages.warning(request, f'⚠️ {duplicates} کلمه تکراری بود و نادیده گرفته شد.')

    return redirect('rank_tracker:project_detail', project_id=project.id)


@login_required
def keyword_delete(request, keyword_id):
    """حذف کلمه کلیدی"""
    keyword = get_object_or_404(RankKeyword, id=keyword_id, project__user=request.user)
    project_id = keyword.project.id

    keyword.delete()
    messages.success(request, f'✅ کلمه کلیدی "{keyword.keyword}" حذف شد.')

    return redirect('rank_tracker:project_detail', project_id=project_id)


@login_required
def update_ranks(request, project_id):
    """آپدیت رتبه‌های پروژه (از طریق Celery Task)"""
    project = get_object_or_404(RankProject, id=project_id, user=request.user)

    if request.method == 'POST':
        # آپدیت سنکرون (برای نسخه اولیه - بعداً Celery اضافه می‌کنیم)
        rank_service = RankService()
        result = rank_service.update_project_ranks(project)

        if result['status'] == 'success':
            messages.success(
                request,
                f"✅ آپدیت انجام شد! موفق: {result['updated']} | ناموفق: {result['failed']}"
            )
        else:
            messages.error(request, f"❌ {result['error']}")

        return redirect('rank_tracker:project_detail', project_id=project.id)

    return redirect('rank_tracker:project_detail', project_id=project.id)


@login_required
def keyword_history_api(request, keyword_id):
    """API برای دریافت تاریخچه رتبه یک کلمه کلیدی (برای Chart)"""
    keyword = get_object_or_404(RankKeyword, id=keyword_id, project__user=request.user)

    # دریافت تاریخچه
    history = keyword.history.order_by('checked_at').all()[:30]  # آخرین 30 رکورد

    data = {
        'keyword': keyword.keyword,
        'dates': [h.checked_at.strftime('%Y-%m-%d') for h in history],
        'ranks': [h.rank if h.rank else None for h in history]
    }

    return JsonResponse(data)


@login_required
def download_sample_excel(request):
    """دانلود فایل نمونه Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Keywords"

    # هدر
    ws['A1'] = 'keyword'
    ws['A1'].font = Font(bold=True, size=12)

    # نمونه داده
    sample_keywords = [
        'خرید گوشی موبایل',
        'قیمت لپ تاپ',
        'بهترین تبلت',
        'خرید آنلاین',
        'فروشگاه اینترنتی'
    ]

    for idx, keyword in enumerate(sample_keywords, start=2):
        ws[f'A{idx}'] = keyword

    # توضیحات
    ws['C1'] = '📝 راهنما:'
    ws['C1'].font = Font(bold=True, size=11, color='0000FF')
    ws['C2'] = '✅ فقط یک ستون با نام "keyword"'
    ws['C3'] = '✅ هر سطر یک کلمه کلیدی'
    ws['C4'] = '✅ حداکثر تا ظرفیت پروژه'
    ws['C5'] = '✅ فرمت: xlsx یا xls'

    # تنظیم عرض ستون‌ها
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['C'].width = 35

    # ذخیره و ارسال
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=keywords_sample.xlsx'
    wb.save(response)

    return response
